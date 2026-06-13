"""
Translate Excel text content from English to Chinese.
Preserves format, structure, and formulas.
"""
import json
import re
import sys
from datetime import date, datetime, time
from pathlib import Path

from deep_translator import MyMemoryTranslator
from openpyxl import load_workbook

TRANSLATOR = GoogleTranslator(source="en", target="zh-CN")
MAX_CHARS = 4500  # Google Translate limit per request


def is_translatable(value):
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return False
    if isinstance(value, (datetime, date, time)):
        return False
    if not isinstance(value, str):
        return False
    s = value.strip()
    if not s:
        return False
    if s.startswith("="):
        return False
    if re.match(r"^[\d\s.,\-+%$€£¥]+$", s):
        return False
    return True


def translate_text(text: str) -> str:
    text = text.strip()
    if not text:
        return text
    if len(text) <= MAX_CHARS:
        return TRANSLATOR.translate(text)
    parts = []
    for i in range(0, len(text), MAX_CHARS):
        chunk = text[i : i + MAX_CHARS]
        parts.append(TRANSLATOR.translate(chunk))
    return "".join(parts)


def translate_excel(input_path: str, output_path: str) -> dict:
    wb = load_workbook(input_path, data_only=False)
    sheets_translated = []
    sheets_skipped = []
    notes = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        translated_count = 0
        for row in ws.iter_rows():
            for cell in row:
                val = cell.value
                if not is_translatable(val):
                    continue
                try:
                    translated = translate_text(str(val))
                    cell.value = translated
                    translated_count += 1
                except Exception as e:
                    notes.append(f"{sheet_name}!{cell.coordinate}: {e}")
        if translated_count > 0:
            sheets_translated.append(sheet_name)
        else:
            sheets_skipped.append(sheet_name)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()

    return {
        "sheets_translated": sheets_translated,
        "sheets_skipped": sheets_skipped,
        "translation_complete": len(sheets_skipped) == 0 or len(sheets_translated) > 0,
        "notes": "; ".join(notes) if notes else "Translation completed successfully.",
    }


def main():
    if len(sys.argv) < 3:
        print("Usage: python translate_excel.py <input.xlsx> <output.xlsx>")
        sys.exit(2)
    input_path = sys.argv[1]
    output_path = sys.argv[2]
    metrics = translate_excel(input_path, output_path)
    metrics_path = Path(output_path).parent / "metrics.json"
    with open(metrics_path, "w", encoding="utf-8") as f:
        json.dump(metrics, f, ensure_ascii=False, indent=2)
    print(f"Saved: {output_path}")
    print(f"Metrics: {metrics_path}")


if __name__ == "__main__":
    main()
()
