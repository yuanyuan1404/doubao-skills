"""
inspect_workbook.py
工作簿结构检查脚本：读取所有 sheet 名称，并为每个 sheet 输出前 15 行原始矩阵。
针对数据处理场景优化：重点关注数据类型、缺失值、合并单元格等信息。
要求：不省略空行；识别并输出合并单元格范围。
"""
import json
import sys
from datetime import date, datetime, time

import openpyxl

MAX_PREVIEW_ROWS = 15

SPECIAL_ROW_KEYWORDS = [
    "合计",
    "总计",
    "小计",
    "汇总",
    "累计",
    "总和",
    "grand total",
    "subtotal",
    "total",
]


def cell_value(cell):
    v = cell.value
    if v is None:
        return None
    if isinstance(v, (datetime, date, time)):
        return v.isoformat()
    if isinstance(v, (int, float)):
        return v
    return v


def detect_special_rows(ws, max_scan_top=200, max_scan_bottom=50, max_results=30):
    max_row = ws.max_row
    max_col = ws.max_column
    rows_to_scan = set()
    for r in range(1, min(max_row, max_scan_top) + 1):
        rows_to_scan.add(r)
    if max_row > 1:
        for r in range(max(1, max_row - max_scan_bottom + 1), max_row + 1):
            rows_to_scan.add(r)

    results = []
    for r in sorted(rows_to_scan):
        if r <= 1:
            continue
        matched = set()
        samples = []
        for c in range(1, max_col + 1):
            v = ws.cell(row=r, column=c).value
            if not isinstance(v, str):
                continue
            s = v.strip()
            if not s:
                continue
            s_lower = s.lower()
            for kw in SPECIAL_ROW_KEYWORDS:
                if kw.isascii():
                    if kw in s_lower:
                        matched.add(kw)
                else:
                    if kw in s:
                        matched.add(kw)
            if matched and len(samples) < 3:
                samples.append({"col": openpyxl.utils.get_column_letter(c), "value": s[:80]})
        if matched:
            results.append(
                {
                    "row_index": r,
                    "matched_keywords": sorted(matched),
                    "sample_cells": samples,
                }
            )
            if len(results) >= max_results:
                break
    return results


def is_empty_header_value(value):
    if value is None:
        return True
    if isinstance(value, str) and value.strip() == "":
        return True
    return False


def row_type_summary(ws, row_idx, max_col):
    counts = {"empty": 0, "string": 0, "numeric": 0, "datetime": 0, "other": 0}
    non_empty_string_count = 0
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_idx, column=c)
        t = get_cell_data_type(cell)
        counts[t] = counts.get(t, 0) + 1
        if t == "string" and isinstance(cell.value, str) and cell.value.strip() != "":
            non_empty_string_count += 1
    counts["non_empty_string"] = non_empty_string_count
    return counts


def suggest_header_row(ws, max_rows_to_check=5):
    max_col = ws.max_column
    max_row = min(max_rows_to_check, ws.max_row)
    candidates = []
    best = None
    for r in range(1, max_row + 1):
        summary = row_type_summary(ws, r, max_col)
        score = summary["non_empty_string"] * 2 + summary["datetime"] - summary["numeric"]
        item = {"row_index": r, "score": score, "summary": summary}
        candidates.append(item)
        if best is None or score > best["score"]:
            best = item
    return {
        "candidates": candidates,
        "suggested": best["row_index"] if best else None,
    }


def get_cell_data_type(cell):
    """获取单元格数据类型"""
    if cell.value is None:
        return "empty"
    if isinstance(cell.value, (int, float)):
        return "numeric"
    if isinstance(cell.value, (datetime, date, time)):
        return "datetime"
    if isinstance(cell.value, str):
        return "string"
    return "other"


def merged_ranges(ws):
    return [str(r) for r in ws.merged_cells.ranges]


def merged_by_row(ws):
    row_map = {}
    for r in ws.merged_cells.ranges:
        for row in range(r.min_row, r.max_row + 1):
            row_map.setdefault(row, []).append(str(r))
    return row_map


def extract_preview(ws):
    preview = {}
    max_col = ws.max_column
    max_row = min(MAX_PREVIEW_ROWS, ws.max_row)
    for r in range(1, max_row + 1):
        row_vals = []
        row_types = []
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            row_vals.append(cell_value(cell))
            row_types.append(get_cell_data_type(cell))
        preview[r] = {
            "values": row_vals,
            "types": row_types
        }
    return preview


def analyze_column(ws, col_idx):
    """分析单列数据特征"""
    col_letter = openpyxl.utils.get_column_letter(col_idx)
    values = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        if cell.value is not None:
            values.append(cell_value(cell))
    
    if not values:
        return {"type": "empty", "non_empty_count": 0}
    
    types_count = {}
    for v in values:
        t = type(v).__name__
        types_count[t] = types_count.get(t, 0) + 1
    
    dominant_type = max(types_count, key=types_count.get)
    
    return {
        "type": dominant_type,
        "non_empty_count": len(values),
        "empty_count": ws.max_row - 1 - len(values),
        "type_distribution": types_count
    }


def build_profile(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    result = {
        "file": path,
        "sheet_names": wb.sheetnames,
        "sheets": [],
    }
    for name in wb.sheetnames:
        ws = wb[name]
        
        col_analysis = {}
        for c in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            header = ws.cell(row=1, column=c).value
            col_analysis[col_letter] = {
                "header": header,
                "analysis": analyze_column(ws, c)
            }

        header_row_1 = []
        empty_header_columns = []
        for c in range(1, ws.max_column + 1):
            col_letter = openpyxl.utils.get_column_letter(c)
            header_value = ws.cell(row=1, column=c).value
            header_row_1.append(header_value)
            if is_empty_header_value(header_value):
                analysis = col_analysis.get(col_letter, {}).get("analysis", {})
                if analysis.get("non_empty_count", 0) > 0:
                    empty_header_columns.append(col_letter)
        
        result["sheets"].append({
            "sheet": name,
            "max_row": ws.max_row,
            "max_col": ws.max_column,
            "preview_rows": extract_preview(ws),
            "merged_cells": merged_ranges(ws),
            "merged_cells_by_row": merged_by_row(ws),
            "header_row_1": header_row_1,
            "empty_header_columns": empty_header_columns,
            "header_row_detection": suggest_header_row(ws),
            "column_analysis": col_analysis,
        })
    return result


def main():
    if len(sys.argv) < 2:
        print("Usage: python scripts/inspect_workbook.py <excel_path> [output_json]")
        sys.exit(2)
    path = sys.argv[1]
    out = sys.argv[2] if len(sys.argv) > 2 else "probe_result.json"
    profile = build_profile(path)
    with open(out, "w", encoding="utf-8") as f:
        json.dump(profile, f, ensure_ascii=False, indent=2)
    print(f"Wrote: {out}")


if __name__ == "__main__":
    main()
