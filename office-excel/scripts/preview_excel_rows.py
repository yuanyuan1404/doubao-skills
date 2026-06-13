#!/usr/bin/env python3
"""Preview the first N rows of an Excel sheet for header inspection."""

import argparse
import json
from datetime import datetime
from typing import Any, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ")
    return str(value)


def _escape_md(value: str) -> str:
    return value.replace("|", "\\|")


def preview_rows(path: str, sheet: Optional[str], rows: int) -> dict:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet] if sheet else workbook.active

    preview: List[List[str]] = []
    max_cols = 0
    for row in worksheet.iter_rows(min_row=1, max_row=rows, values_only=True):
        values = [_stringify(cell) for cell in row]
        max_cols = max(max_cols, len(values))
        preview.append(values)

    column_letters = [get_column_letter(i + 1) for i in range(max_cols)]
    non_empty_columns = []
    for col_index in range(max_cols):
        if any(r[col_index] if col_index < len(r) else "" for r in preview):
            non_empty_columns.append(get_column_letter(col_index + 1))

    return {
        "sheet": worksheet.title,
        "rows": [
            {
                "row_index": i + 1,
                "values": r + [""] * (max_cols - len(r)),
            }
            for i, r in enumerate(preview)
        ],
        "max_columns": max_cols,
        "column_letters": column_letters,
        "non_empty_columns": non_empty_columns,
    }


def to_markdown(data: dict) -> str:
    column_letters = data["column_letters"]
    header = ["row"] + column_letters
    lines = []
    lines.append(f"Sheet: {data['sheet']}")
    lines.append(f"Rows previewed: {len(data['rows'])}")
    lines.append(
        "Non-empty columns: "
        + (", ".join(data["non_empty_columns"]) if data["non_empty_columns"] else "None")
    )
    lines.append("")

    header_row = "| " + " | ".join(header) + " |"
    separator = "| " + " | ".join(["---"] * len(header)) + " |"
    lines.append(header_row)
    lines.append(separator)

    for row in data["rows"]:
        values = [_escape_md(v) for v in row["values"]]
        line = "| " + " | ".join([str(row["row_index"])] + values) + " |"
        lines.append(line)

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Preview the first N rows of an Excel sheet for header inspection."
    )
    parser.add_argument("path", help="Path to .xlsx file")
    parser.add_argument("--sheet", help="Sheet name (default: active sheet)")
    parser.add_argument("--rows", type=int, default=15, help="Number of rows to preview")
    parser.add_argument(
        "--output",
        choices=["markdown", "json"],
        default="markdown",
        help="Output format",
    )
    args = parser.parse_args()

    data = preview_rows(args.path, args.sheet, args.rows)
    if args.output == "json":
        print(json.dumps(data, ensure_ascii=True, indent=2))
    else:
        print(to_markdown(data))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
